from os import path
from PyQt5 import QtGui, QtWidgets, QtCore
from base import from_tuple_to_patients_table
from ui.excel_window import Ui_ExcelWindow
from static import set_text, set_title_font
from PyQt5.QtWidgets import QGridLayout, QTableWidgetItem, QTableWidget, QAbstractItemView, QLabel
import datetime
from PyQt5.QtCore import Qt
import openpyxl
from error_window import ErrorWindow
from static import validator


class ExcelWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(ExcelWindow, self).__init__()
        self.setFixedSize(1172, 705)
        # Инициализация окон
        self.ui = Ui_ExcelWindow()
        self.ui_2 = ErrorWindow()
        self.ui.setupUi(self)
        # Пути до папок
        self.template_dir = path.join(path.dirname(__file__), 'templates')
        self.img_dir = path.join(path.dirname(__file__), 'img')
        # Иконка окна
        self.setWindowIcon(QtGui.QIcon(path.join(self.img_dir, 'gosuslugi_5.png')))
        # Оглавление окна
        self.setWindowTitle('Перенос информации о пациентах из Excel в базу данных')
        # Переменные для данных из экселя
        self.table = QTableWidget(self)
        self.grid_layout = QGridLayout()
        self.headers = []
        self.data = {}
        self.init_handlers()
        # Список для добавления выделенных ячеек
        self.select_list = []
        # Текущая дата
        self.datetime_now = datetime.datetime.now()
        # Валидация
        validator(self.ui.lineEdit, '[0-9]{10}')
        validator(self.ui.lineEdit_2, '[0-9]{10}')
        # Текст по окну
        set_text(self.ui.label, 'С какой строки:')
        set_text(self.ui.label_2, 'По какую строку:')
        set_text(self.ui.label_3, 'Вбейте нужные значения в форму,'
                                  'либо выделите нужные строки левой кн. мыши в любом месте')
        set_title_font(self.ui.label_3)
        set_text(self.ui.pushButton, 'Применить')
        self.ui.pushButton.setStyleSheet("""
                                           background-color: #b2edbf;
                                           """)
        set_text(self.ui.pushButton_2, 'Отмена')
        self.ui.pushButton_2.setStyleSheet("""
                                             background-color: #f7c8c8;
                                             """)

    @QtCore.pyqtSlot(QtCore.QItemSelection, QtCore.QItemSelection)
    def on_selection_changed(self, selected, deselected):
        # Добавление номера выделенной строки в список self.select_list в реальном времени по вызову декоратора
        for ix in selected.indexes():
            self.select_list.append(ix.row())
            # Плюс 1 для корректного отображения пользователю
            first_cell = int(self.select_list[0]) + 1
            last_cell = int(self.select_list[len(self.select_list) - 1]) + 1
            # Если выделение снизу -> вверх
            if first_cell > last_cell:
                first_cell = int(self.select_list[len(self.select_list) - 1]) + 1
                last_cell = int(self.select_list[0]) + 1
            # Добавление в lineEdit
            self.ui.lineEdit.setText(str(first_cell))
            self.ui.lineEdit_2.setText(str(last_cell))

        # Удаление номера выделенной строки после отмены выделения
        for ix in deselected.indexes():
            self.select_list.remove(ix.row())  # Удаление всех выделенных строк из списка
            # Приравнивание двух lineEdit для одной выделенной ячейки
            first_cell = int(self.select_list[0]) + 1
            self.ui.lineEdit.setText(str(first_cell))

    # Подключение кнопок
    def init_handlers(self):
        self.ui.pushButton.clicked.connect(self.excel_to_bd)
        self.ui.pushButton_2.clicked.connect(self.close_window)
        self.table.selectionModel().selectionChanged.connect(self.on_selection_changed)

    # Закрытие окна
    def close_window(self):
        self.close()

    # Отображение окна ошибок
    def show_error_window(self, error):
        label = self.ui_2.findChildren(QLabel)

        for item in label:
            item.setText(error)

        self.ui_2.show()

    # Заполнение таблицы
    def refresh(self):
        # Инициализация таблицы
        self.ui.tableWidget.setLayout(self.grid_layout)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        # Функция на чтение экселя
        self.read_excel()
        # Кол-во строк и колонок для отрисовки таблицы
        self.table.setColumnCount(len(self.headers))
        self.table.setRowCount(len(self.data))
        # Заполнение таблицы
        self.fill_table()
        self.grid_layout.addWidget(self.table, 0, 0)  # Добавляем таблицу в сетку

    # Забирает данные из экселя
    def read_excel(self):
        # Открытие экселя в папке
        excel = openpyxl.load_workbook(filename=path.join(self.template_dir, 'template.xlsx'), data_only=True)
        sheet = excel[excel.sheetnames[0]]  # Номер листа
        max_row = sheet.max_row  # Максимальное количество строк
        max_col = sheet.max_column  # Максимальное количество колонок

        # Перебор для header
        for row in range(1, 2):
            for col in range(max_col):
                # Нумерация в openpyxl начинается с 1, 1 а не 0,0
                self.headers.append(sheet.cell(row + 1, col + 1).value)

        # Перебор всех ячеек после header
        for row in range(2, max_row):
            excel_list = []  # Список для value словаря
            for col in range(max_col):
                cell = sheet.cell(row + 1, col + 1).value  # Нумерация в openpyxl начинается с 1, 1 а не 0,0
                if type(cell) == datetime.datetime:  # Проверка на datetime
                    excel_list.append(cell.strftime('%Y-%m-%d'))
                elif cell is None:  # Проверка на пустую ячейку
                    excel_list.append('')
                else:
                    excel_list.append(str(cell))

            self.data[int(row)] = excel_list  # Добавление значений ячеек в словарь

    # Наполнение таблицы
    def fill_table(self):
        # Устанавливаем заголовки таблицы
        self.table.setHorizontalHeaderLabels(self.headers)

        # Устанавливаем выравнивание на заголовки
        for el in range(len(self.headers)):
            self.table.horizontalHeaderItem(el).setTextAlignment(Qt.AlignHCenter)
        # Заполняем строки
        for keys, values in self.data.items():
            for el in range(len(self.data[keys])):
                self.table.setItem(keys - 2, el, QTableWidgetItem(self.data[keys][el]))

        # делаем ресайз колонок по содержимому
        self.table.resizeColumnsToContents()

    def excel_to_bd(self):
        try:
            # Доп. данные для БД
            date = self.datetime_now.strftime("%d-%m-%Y")  # Текущая дата
            success = 0  # Сведения об отправке

            # Текущие значения в полях lineEdit
            start_row = self.ui.lineEdit.text()
            finish_row = self.ui.lineEdit_2.text()
            if int(start_row) > int(finish_row):
                self.show_error_window(f'Значение ОТ больше значения ДО')
                return

            data = {}  # Словарь для отправки в базу

            # Перебор строк
            for row in range(int(start_row) - 1, int(finish_row)):
                patient_list = []  # Список для сбора данных с ячеек таблицы
                # Перебор колонок
                for column in range(len(self.headers)):
                    cell_text = str(self.table.item(row, column).text())
                    patient_list.append(cell_text)
                # Добавление доп. данных для БД
                patient_list.insert(0, date)
                patient_list.append(success)
                data[row] = patient_list

            # Перебор по словарю и добавление в БД
            for value in data.values():
                from_tuple_to_patients_table(tuple(value))

            # Открытие окна об успешном выполнении
            self.show_error_window(f'Сведения успешно внесены в базу данных.\n'
                                   f'Количество пациентов: {int(finish_row) - int(start_row) + 1}')
            self.close_window()
            # Сброс списка хедеров, чтобы при повторном открытии он обновлялся, а не дополнялся
            self.headers = []
        except ValueError:
            self.show_error_window(f'Введите значения в поля ввода\nили выделите нужные ячейки левой кн.мыши')
        except Exception as error:
            self.show_error_window(f'{error}')  # Сбор ошибок в окно ошибок
