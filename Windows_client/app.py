import os
import re
import sys
import datetime
import openpyxl
import configparser
from os import path
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import *
from docxtpl import DocxTemplate
from PyQt5 import QtWidgets, QtGui
from about_us import AboutUsWindow
from error_window import ErrorWindow
from config_window import ConfigWindow
from ui.main_window import Ui_MainWindow
from openpyxl.utils import get_column_letter
from static import open_readme, lock_form, set_text


class main_window(QtWidgets.QMainWindow):
    def __init__(self):
        super(main_window, self).__init__()
        self.setFixedSize(443, 491)
        # Инициализация окон
        self.ui = Ui_MainWindow()
        self.ui_2 = ConfigWindow()
        self.ui_3 = AboutUsWindow()
        self.ui_4 = ErrorWindow()
        self.ui.setupUi(self)
        # Подключение других окон
        self.init_handlers()
        # Дата в реальном времени
        self.datetime_now = datetime.datetime.now()
        # Расширения для шаблонов
        self.excel_extends = [".xlsm", ".xlsx", ".xls"]
        self.word_extends = [".doc", ".docx"]
        # Пути до папок
        self.template_dir = path.join(path.dirname(__file__), 'templates')
        self.config_dir = path.join(path.dirname(__file__), 'config')
        self.result_dir = path.join(path.dirname(__file__), 'result')
        # Обьект класса на чтение конфигов
        self.config = configparser.RawConfigParser()
        self.config.read(path.join(self.config_dir, 'config.ini'))
        # Открытие файла, не зная его расширения
        self.open_file_without_extend()
        self.read_excel_file()
        # Прогрессбар
        self.ui.progressBar.hide()
        # Текст по приложению
        self.windows_text()

    # Обработка нажатия для октрытия сторонних окон
    def init_handlers(self):
        self.ui.toolButton.clicked.connect(self.get_excel_file)
        self.ui.toolButton_2.clicked.connect(self.get_word_file)
        self.ui.toolButton_3.clicked.connect(self.show_config_window)
        self.ui.toolButton_4.clicked.connect(self.show_about_us)
        self.ui.toolButton_5.clicked.connect(open_readme)
        self.ui.pushButton.clicked.connect(self.excel_to_word)
        # Радио кнопки
        self.ui.radioButton.setChecked(True)
        self.ui.comboBox_2.setEnabled(False)
        self.ui.radioButton_2.clicked.connect(self.set_enabled)
        self.ui.radioButton.clicked.connect(self.set_not_enabled)

    # Оформление окна
    def windows_text(self):
        # Иконка для приложения
        self.setWindowIcon(QtGui.QIcon(path.join(self.config_dir, 'logo.ico')))
        # Оглавление приложения
        self.setWindowTitle('Эксель -> Ворд')
        # Текст по окну
        set_text(self.ui.label, 'Место нахождения Excel шаблона')
        set_text(self.ui.label_2, 'С какой строки: ')
        set_text(self.ui.label_4, 'Место нахождения Word шаблона')
        set_text(self.ui.label_3, 'Нажмите кнопку для формирования документа')
        set_text(self.ui.label_5, 'По какую строку: ')
        self.ui.label_3.setAlignment(Qt.AlignCenter)
        set_text(self.ui.pushButton, 'Сформировать')
        self.ui.pushButton.setStyleSheet("""
                                           background-color: #d2dffa
                                           """)
        set_text(self.ui.toolButton_3, 'Конфигурации')
        set_text(self.ui.toolButton_4, 'О проекте')
        set_text(self.ui.toolButton_5, 'Руководство')
        set_text(self.ui.radioButton, 'Одна строка')
        set_text(self.ui.radioButton_2, 'Несколько строк')
        # Ограничение на 25 отображаемых элементов списка combobox
        self.ui.comboBox.setStyleSheet("QComboBox { combobox-popup: 0; }")  # Нужно для стиля Fusion
        self.ui.comboBox_2.setStyleSheet("QComboBox { combobox-popup: 0; }")  # Так как список combobox нередактируемый
        self.ui.comboBox.setMaxVisibleItems(25)
        self.ui.comboBox_2.setMaxVisibleItems(25)

    # Открытие окна конфигов
    def show_config_window(self):
        self.ui_2.show()

    # Открытие окна О нас
    def show_about_us(self):
        self.ui_3.show()

    # Открытие окна событий
    def show_error_window(self, error):
        label = self.ui_4.findChildren(QLabel)

        for item in label:
            item.setText(error)

        self.ui_4.show()

    def set_enabled(self):
        self.ui.comboBox_2.setEnabled(True)

    def set_not_enabled(self):
        self.ui.comboBox_2.setEnabled(False)

    # Открытие файла без указанного расширения
    def open_file_without_extend(self):
        for ext in self.excel_extends:
            if path.isfile(path.join(self.template_dir, 'excel_template') + ext):
                set_text(self.ui.lineEdit, path.join(self.template_dir, 'excel_template') + ext)
                lock_form(self.ui.lineEdit)
        for ext in self.word_extends:
            if path.isfile(path.join(self.template_dir, 'word_template') + ext):
                set_text(self.ui.lineEdit_2, path.join(self.template_dir, 'word_template') + ext)
                lock_form(self.ui.lineEdit_2)

    # Диалоговое окно на выбор Эксель файла
    def get_excel_file(self):
        filename, filetype = QFileDialog.getOpenFileName(self, "Выбрать файл", f"{self.template_dir}",
                                                         "Excel Files(*.xlsx);;"
                                                         "Old-excel Files(*.xls);;"
                                                         "Excel-macros Files(*.xlsm);;"
                                                         "All Files(*)")
        # Обработка того, что пользователь может и не выбрать файл
        if filename:
            set_text(self.ui.lineEdit, filename)
            lock_form(self.ui.lineEdit)
            self.read_excel_file()

    # Диалоговое окно на выбор Ворд файла
    def get_word_file(self):
        filename, filetype = QFileDialog.getOpenFileName(self, "Выбрать файл", f"{self.template_dir}",
                                                         "Word Files(*.docx);;"
                                                         "Old-word Files(*.doc);;"
                                                         "All Files(*)")
        # Обработка того, что пользователь может и не выбрать файл
        if filename:
            set_text(self.ui.lineEdit_2, filename)
            lock_form(self.ui.lineEdit_2)

    # Чтение эксель файла и передача списка ячеек в comboBox
    def read_excel_file(self):
        excel = openpyxl.load_workbook(filename=self.ui.lineEdit.text(), data_only=True)
        # Список для вывода значений в comboBox
        excel_list = []
        self.ui.comboBox.clear()
        self.ui.comboBox_2.clear()
        # Перебор конфига
        for section in self.config.sections():
            if self.config.has_section('combobox'):
                if self.config.has_option(section, 'combo_cell') and self.config.has_option(section, 'combo_sheet'):
                    combobox_row = self.config.get(section, 'combo_cell')
                    combobox_sheet = self.config.get(section, 'combo_sheet')
                    sheet = excel[excel.sheetnames[int(combobox_sheet) - 1]]
                    row_max = len(sheet[combobox_row])
                    # Последняя цифра ({rows}) в выводе нужна для указывания по какому номеру строки работать циклу
                    for rows in range(1, row_max):
                        if sheet[f'{combobox_row}{rows}'].value is not None:
                            excel_list.append(f"{str(rows).zfill(4)} | {sheet[f'{combobox_row}{rows}'].value}")
        # Добавление в comboBox
        for el in excel_list:
            length = 75
            if len(el) > length:
                self.ui.comboBox.addItem(f'{el[:length]}...')
                self.ui.comboBox_2.addItem(f'{el[:length]}...')
            else:
                self.ui.comboBox.addItem(el)
                self.ui.comboBox_2.addItem(el)

    # Цикл для поиска по Эксель файлу
    def excel_search(self, find_value, find_sheet):
        excel = openpyxl.load_workbook(filename=self.ui.lineEdit.text(), data_only=True)
        sheet = excel[excel.sheetnames[int(find_sheet)-1]]

        column_min = 1
        row_min = 1
        row_max = sheet.max_row

        # Флаг для выхода из внешнего цикла
        while_flag = True
        while while_flag:
            row_counter = row_min
            # Проходит по колонке от 1 до max
            while row_counter <= row_max:
                word_column = str(get_column_letter(column_min))
                data_from_cell = str(sheet[word_column + str(row_counter)].value)
                result = re.search(find_value, data_from_cell)
                # Ищем только первое вхождение
                if result:
                    new_column = column_min + 1  # Получает значение справа на 1 ячейку от найденной ячейки
                    return_value = sheet[f'{get_column_letter(new_column)}{row_counter}'].value
                    return return_value
                row_counter = int(row_counter) + 1

            column_min = int(column_min) + 1

    # Метод для переноса ячеек в шаблон Ворд
    def excel_to_word(self):
        # Прогрессбар каунтер
        self.ui.progressBar.show()
        progress = 0

        excel = openpyxl.load_workbook(filename=self.ui.lineEdit.text(), data_only=True)
        # Ищет именно по comboBox
        row = self.ui.comboBox.currentText()  # Последняя цифра в комбобоксе - номер строки
        row = row[:4]
        row_2 = row

        if self.ui.radioButton_2.isChecked():
            row_2 = self.ui.comboBox_2.currentText()
            row_2 = row_2[:4]

        if row_2 < row:
            self.show_error_window('Значение "ДО" больше значения "ПОСЛЕ"')
            self.ui.progressBar.setValue(0)
            self.ui.progressBar.hide()
            return

        for rows in range(int(row), int(row_2) + 1):
            # Два словаря нужны для проверки ячейки на валидность даты
            merge_dict_before = {}
            merge_dict_after = {}
            # Чтение конфига config.ini
            for section in self.config.sections():
                if self.config.has_option(section, 'cell') and self.config.has_option(section, 'sheet'):
                    cell = self.config.get(section, 'cell')
                    excel_sheet = self.config.get(section, 'sheet')
                    sheet = excel[excel.sheetnames[int(excel_sheet) - 1]]

                    # Валидация на явно указанную ячейку (А1)
                    if any(map(str.isdigit, cell)):  # Проверка на цифру в строке
                        merge_dict_before[f'{section}'] = sheet[f'{cell}'].value
                    else:
                        merge_dict_before[f'{section}'] = sheet[f'{cell}{rows}'].value

                if self.config.has_option(section, 'date'):
                    merge_dict_before[f'datetime'] = self.datetime_now

                # Проверка на поиск по другому листу
                if self.config.has_option(section, 'find_sheet'):
                    find_sheet = self.config.get(section, 'find_sheet')
                    find_cell = merge_dict_before[f'{section}']
                    new_value = self.excel_search(str(find_cell), find_sheet)
                    merge_dict_before[f'{section}find'] = new_value

            # Валидация на тип "datetime.datetime" и перенос данных во второй словарь
            for key, value in merge_dict_before.items():
                if value is None:
                    merge_dict_after[key] = '-'
                elif type(value) is datetime.datetime:
                    merge_dict_after[key] = value.strftime('%H:%M:%S %d.%m.%Y')
                    merge_dict_after[f'{key}hour'] = value.strftime('%H')
                    merge_dict_after[f'{key}minute'] = value.strftime('%M')
                    merge_dict_after[f'{key}time'] = value.strftime('%H:%M:%S')
                    merge_dict_after[f'{key}day'] = value.strftime('%d')
                    merge_dict_after[f'{key}month'] = value.strftime('%m')
                    merge_dict_after[f'{key}year'] = value.strftime('%Y')[2:]
                    merge_dict_after[f'{key}date'] = value.strftime('%d.%m.%Y')
                else:
                    merge_dict_after[key] = value

            # Название конечного документа в виде даты и времени создания
            date_format = self.datetime_now.strftime('%d-%m-%Y')
            # Рендер в Ворд шаблон
            doc = DocxTemplate(self.ui.lineEdit_2.text())
            doc.render(merge_dict_after)
            # Сохранение и открытие файла
            doc.save(f"{path.join(self.result_dir, f'{date_format} - {rows}.docx')}")
            # Добавляем значение в прогрессбар
            progress += int(100 / (int(row_2) - int(row) + 1))
            self.ui.progressBar.setValue(progress)

        # Открытие папки result после завершения цикла
        os.startfile(self.result_dir)
        # Обнуляем прогрессбар
        self.ui.progressBar.setValue(0)
        self.ui.progressBar.hide()


os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"  # Адаптирование размера приложения для высокого разрешения
app = QtWidgets.QApplication([])
app.setStyle('Fusion')  # Оформление всего приложения
application = main_window()
application.show()
sys.exit(app.exec())
